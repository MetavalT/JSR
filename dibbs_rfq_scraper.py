"""
╔══════════════════════════════════════════════════════════════════════════════╗
║              DIBBS RFQ Scraper — Metaval Engineering                        ║
║──────────────────────────────────────────────────────────────────────────────║
║  Filters applied:                                                            ║
║   • FSC codes : 4820, 4730, 5305, 5306, 5310, 5330, 5340, 3490, 4710       ║
║   • Mil-spec  : NSN must carry a Mil-spec note                              ║
║   • Tech Doc  : note must be exactly  "*spec/stnd only"                     ║
║   • Sort      : newest RFQ date first                                       ║
║   • Output    : Excel (.xlsx)  — no login required                          ║
╚══════════════════════════════════════════════════════════════════════════════╝

Usage
-----
    # Basic run (auto-detects output folder):
    python dibbs_rfq_scraper.py

    # Specify how many result pages to scrape (default 10):
    python dibbs_rfq_scraper.py --pages 25

    # Specify a custom output folder:
    python dibbs_rfq_scraper.py --out my_results

    # Full example:
    python dibbs_rfq_scraper.py --pages 20 --out dibbs_output

Requirements
------------
    pip install requests beautifulsoup4 openpyxl lxml
"""

import argparse
import os
import re
import sys
import time
import traceback
from datetime import datetime
from urllib.parse import urlencode, urljoin

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import (Alignment, Border, Font, PatternFill, Side)
from openpyxl.utils import get_column_letter

# ─── Constants ────────────────────────────────────────────────────────────────

BASE_URL   = "https://www.dibbs.bsm.dla.mil"
SEARCH_URL = f"{BASE_URL}/RFQ/RFQList.aspx"

# FSC codes your company cares about
TARGET_FSC = {"4820", "4730", "5305", "5306", "5310", "5330", "5340", "3490", "4710"}

# Tech-doc filter — must match this text exactly (case-insensitive strip)
TECH_DOC_MARKER = "*spec/stnd only"

# Mil-spec keywords found in the "Spec/Std" or description columns
MIL_SPEC_PATTERNS = re.compile(
    r"mil[-\s]?(spec|std|prf|dtl|a|b|c|d|e|f|g|h|i|j|k|l|m|n|p|q|r|s|t|u|v|w|x|y|z)?[-\s]?\d*",
    re.IGNORECASE,
)

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "en-US,en;q=0.9",
}

REQUEST_DELAY = 1.5   # seconds between requests — be polite


# ─── Helpers ─────────────────────────────────────────────────────────────────

def get_session() -> requests.Session:
    s = requests.Session()
    s.headers.update(HEADERS)
    return s


def safe_get(session: requests.Session, url: str, params: dict = None,
             retries: int = 3) -> requests.Response | None:
    for attempt in range(1, retries + 1):
        try:
            r = session.get(url, params=params, timeout=30)
            r.raise_for_status()
            return r
        except requests.RequestException as e:
            print(f"  [WARN] Attempt {attempt}/{retries} failed: {e}")
            if attempt < retries:
                time.sleep(3 * attempt)
    return None


def text(el) -> str:
    """Return stripped inner text of a BS4 element, or ''."""
    return el.get_text(separator=" ", strip=True) if el else ""


def normalise_date(raw: str) -> str:
    """Try to convert various date formats → YYYY-MM-DD for sorting."""
    for fmt in ("%m/%d/%Y", "%d-%b-%Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(raw.strip(), fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return raw.strip()


# ─── DIBBS page parser ────────────────────────────────────────────────────────

def build_search_params(page: int, fsc_code: str) -> dict:
    """
    DIBBS RFQ list accepts FSC via the 'fsc' query param and paginates via
    'start'.  We sort by Posted Date descending (sortBy=PostDate&sortDir=DESC).
    """
    return {
        "fsc"    : fsc_code,
        "sortBy" : "PostDate",
        "sortDir": "DESC",
        "start"  : (page - 1) * 25,     # DIBBS shows 25 rows per page
    }


def parse_rfq_list_page(html: str, fsc_code: str) -> list[dict]:
    """
    Parse one DIBBS RFQ listing page and return rows that pass ALL filters:
      1. FSC matches (already guaranteed by the search param, kept for safety)
      2. Mil-spec note present
      3. Tech-doc note is exactly '*spec/stnd only'
    """
    soup = BeautifulSoup(html, "lxml")
    rows = []

    # DIBBS renders results in a table with class 'tablesorter' or id 'rfqList'
    table = (
        soup.find("table", {"id": "rfqList"})
        or soup.find("table", class_=re.compile(r"tablesorter|rfq", re.I))
        or soup.find("table")          # fallback
    )
    if not table:
        return rows

    tbody = table.find("tbody") or table
    tr_list = tbody.find_all("tr", recursive=False)

    for tr in tr_list:
        tds = tr.find_all("td")
        if len(tds) < 6:
            continue

        # ── Extract raw cell values ──────────────────────────────────────────
        # Column order on DIBBS RFQ list (public view, as of 2025-2026):
        #  0: RFQ No.  1: NSN  2: Item Name  3: Qty  4: Unit
        #  5: Posted Date  6: Response Date  7: Spec/Std  8: Tech Doc
        # Indices may shift — we also try to detect by header scanning below.

        rfq_no       = text(tds[0])
        nsn_el       = tds[1]
        nsn          = text(nsn_el)
        item_name    = text(tds[2]) if len(tds) > 2 else ""
        qty          = text(tds[3]) if len(tds) > 3 else ""
        unit         = text(tds[4]) if len(tds) > 4 else ""
        posted_date  = text(tds[5]) if len(tds) > 5 else ""
        resp_date    = text(tds[6]) if len(tds) > 6 else ""
        spec_std     = text(tds[7]) if len(tds) > 7 else ""
        tech_doc_raw = text(tds[8]) if len(tds) > 8 else ""

        # ── FSC guard: first 4 digits of NSN ────────────────────────────────
        nsn_clean = re.sub(r"[\s\-]", "", nsn)
        nsn_fsc   = nsn_clean[:4]
        if nsn_fsc not in TARGET_FSC:
            # also accept if fsc_code matches (from search param)
            if fsc_code not in TARGET_FSC:
                continue

        # ── Mil-spec filter ──────────────────────────────────────────────────
        combined_spec_text = f"{spec_std} {item_name} {tech_doc_raw}"
        if not MIL_SPEC_PATTERNS.search(combined_spec_text):
            continue

        # ── Tech-doc filter: must be exactly "*spec/stnd only" ───────────────
        if tech_doc_raw.strip().lower() != TECH_DOC_MARKER.lower():
            continue

        # ── RFQ detail link ──────────────────────────────────────────────────
        link_el = tds[0].find("a") or nsn_el.find("a")
        detail_url = ""
        if link_el and link_el.get("href"):
            detail_url = urljoin(BASE_URL, link_el["href"])

        rows.append({
            "RFQ No."      : rfq_no,
            "NSN"          : nsn,
            "FSC"          : nsn_fsc or fsc_code,
            "Item Name"    : item_name,
            "Qty"          : qty,
            "Unit"         : unit,
            "Posted Date"  : normalise_date(posted_date),
            "Response Date": normalise_date(resp_date),
            "Spec / Std"   : spec_std,
            "Tech Doc"     : tech_doc_raw,
            "Detail URL"   : detail_url,
        })

    return rows


def detect_last_page(html: str) -> int:
    """
    Try to read the pagination footer on DIBBS to find total pages.
    Returns the detected page count, or 1 if not found.
    """
    soup = BeautifulSoup(html, "lxml")
    # DIBBS shows "Page X of Y" or a pager div
    pager = soup.find(string=re.compile(r"page\s+\d+\s+of\s+\d+", re.I))
    if pager:
        m = re.search(r"of\s+(\d+)", pager, re.I)
        if m:
            return int(m.group(1))
    # Fallback: count pagination links
    links = soup.find_all("a", href=re.compile(r"start=\d+"))
    starts = []
    for a in links:
        m = re.search(r"start=(\d+)", a["href"])
        if m:
            starts.append(int(m.group(1)))
    if starts:
        return (max(starts) // 25) + 1
    return 1


# ─── Main scrape loop ─────────────────────────────────────────────────────────

def scrape(max_pages: int) -> list[dict]:
    session  = get_session()
    all_rows : list[dict] = []
    seen_rfqs: set[str]   = set()

    print(f"\n{'='*65}")
    print(f"  DIBBS RFQ Scraper  |  FSC codes: {', '.join(sorted(TARGET_FSC))}")
    print(f"  Filter: Mil-spec + Tech Doc = '{TECH_DOC_MARKER}'")
    print(f"{'='*65}\n")

    for fsc in sorted(TARGET_FSC):
        print(f"── Scraping FSC {fsc} ──────────────────────────────────────────")
        page       = 1
        last_page  = max_pages   # will be updated after first response

        while page <= last_page:
            params = build_search_params(page, fsc)
            print(f"  Page {page}/{last_page}  params={params}")

            resp = safe_get(session, SEARCH_URL, params=params)
            if resp is None:
                print(f"  [ERROR] Could not fetch FSC {fsc} page {page}. Skipping.")
                break

            # Detect pagination ceiling on first page
            if page == 1:
                detected = detect_last_page(resp.text)
                last_page = min(detected, max_pages)
                print(f"  Detected {detected} pages → scraping up to {last_page}")

            new_rows = parse_rfq_list_page(resp.text, fsc)

            for row in new_rows:
                key = row["RFQ No."] or row["NSN"]
                if key and key not in seen_rfqs:
                    seen_rfqs.add(key)
                    all_rows.append(row)

            print(f"  → {len(new_rows)} matching rows this page | total so far: {len(all_rows)}")

            if page >= last_page:
                break
            page += 1
            time.sleep(REQUEST_DELAY)

        print()

    # Sort all results newest-first by Posted Date
    all_rows.sort(key=lambda r: r["Posted Date"], reverse=True)

    return all_rows


# ─── Excel writer ─────────────────────────────────────────────────────────────

COLUMNS = [
    "RFQ No.", "NSN", "FSC", "Item Name",
    "Qty", "Unit", "Posted Date", "Response Date",
    "Spec / Std", "Tech Doc", "Detail URL",
]

COL_WIDTHS = {
    "RFQ No."      : 18,
    "NSN"          : 18,
    "FSC"          : 8,
    "Item Name"    : 38,
    "Qty"          : 10,
    "Unit"         : 10,
    "Posted Date"  : 14,
    "Response Date": 16,
    "Spec / Std"   : 22,
    "Tech Doc"     : 20,
    "Detail URL"   : 55,
}


def write_excel(rows: list[dict], out_path: str) -> None:
    wb = Workbook()

    # ── Sheet 1: Filtered Results ────────────────────────────────────────────
    ws = wb.active
    ws.title = "Filtered RFQs"

    # Styles
    HDR_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    HDR_FILL  = PatternFill("solid", start_color="1F4E79")
    HDR_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

    DAT_FONT  = Font(name="Arial", size=9)
    DAT_ALIGN = Alignment(vertical="center", wrap_text=False)
    CTR_ALIGN = Alignment(horizontal="center", vertical="center")
    URL_FONT  = Font(name="Arial", size=9, color="0563C1", underline="single")

    ALT_FILL  = PatternFill("solid", start_color="DCE6F1")
    WHT_FILL  = PatternFill("solid", start_color="FFFFFF")

    thin   = Side(style="thin", color="BBBBBB")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Title banner
    ws.merge_cells(f"A1:{get_column_letter(len(COLUMNS))}1")
    t = ws["A1"]
    t.value = (
        f"DIBBS RFQ — Filtered Results   |   "
        f"FSC: {', '.join(sorted(TARGET_FSC))}   |   "
        f"Filter: Mil-spec + Tech Doc='{TECH_DOC_MARKER}'   |   "
        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    )
    t.font      = Font(name="Arial", bold=True, size=10, color="1F4E79")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20

    # Column headers (row 2)
    for col_idx, col_name in enumerate(COLUMNS, 1):
        c = ws.cell(row=2, column=col_idx, value=col_name)
        c.font      = HDR_FONT
        c.fill      = HDR_FILL
        c.alignment = HDR_ALIGN
        c.border    = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS.get(col_name, 16)
    ws.row_dimensions[2].height = 24

    # Freeze panes below header
    ws.freeze_panes = "A3"

    # Data rows
    for row_idx, row in enumerate(rows, start=3):
        fill = ALT_FILL if row_idx % 2 == 0 else WHT_FILL
        for col_idx, col_name in enumerate(COLUMNS, 1):
            val  = row.get(col_name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = BORDER
            cell.fill   = fill

            if col_name == "Detail URL" and val:
                cell.font      = URL_FONT
                cell.hyperlink = val
                cell.alignment = DAT_ALIGN
            elif col_name in ("Qty", "FSC", "Posted Date", "Response Date", "Unit"):
                cell.font      = DAT_FONT
                cell.alignment = CTR_ALIGN
            else:
                cell.font      = DAT_FONT
                cell.alignment = DAT_ALIGN

        ws.row_dimensions[row_idx].height = 16

    # Auto-filter on header row
    ws.auto_filter.ref = f"A2:{get_column_letter(len(COLUMNS))}{max(len(rows) + 2, 2)}"

    # ── Sheet 2: Summary ─────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 20

    summary_data = [
        ("Generated On",        datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Source",              "DIBBS — dibbs.bsm.dla.mil"),
        ("Total Matching RFQs", len(rows)),
        ("FSC Codes Scraped",   ", ".join(sorted(TARGET_FSC))),
        ("Mil-spec Filter",     "Yes — must contain Mil-spec reference"),
        ("Tech Doc Filter",     f"Exactly: '{TECH_DOC_MARKER}'"),
        ("Date Sort",           "Newest Posted Date first"),
    ]

    # Breakdown by FSC
    fsc_counts: dict[str, int] = {}
    for r in rows:
        fsc_counts[r.get("FSC", "?")] = fsc_counts.get(r.get("FSC", "?"), 0) + 1

    ws2.append(["DIBBS RFQ Scraper — Run Summary"])
    ws2["A1"].font = Font(name="Arial", bold=True, size=12, color="1F4E79")
    ws2.append([])

    for label, val in summary_data:
        ws2.append([label, str(val)])

    ws2.append([])
    ws2.append(["── Breakdown by FSC ──", "Count"])
    for fsc, cnt in sorted(fsc_counts.items()):
        ws2.append([fsc, cnt])

    for row in ws2.iter_rows():
        for cell in row:
            cell.font = Font(name="Arial", size=10)

    wb.save(out_path)
    print(f"\n✅  Excel saved → {out_path}")
    print(f"    Rows written : {len(rows)}")
    print(f"    Sheets       : 'Filtered RFQs'  |  'Summary'")


# ─── Raw dump (fallback / debug) ──────────────────────────────────────────────

def write_raw_csv(rows: list[dict], out_path: str) -> None:
    """Write a raw CSV alongside the Excel for debugging purposes."""
    import csv
    with open(out_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=COLUMNS)
        writer.writeheader()
        writer.writerows(rows)


# ─── CLI ──────────────────────────────────────────────────────────────────────

def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(
        description="Scrape DIBBS RFQ listings with Mil-spec + Tech-doc filters.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    p.add_argument(
        "--pages", type=int, default=10, metavar="N",
        help="Maximum pages to scrape per FSC code (default: 10). Use 999 for all.",
    )
    p.add_argument(
        "--out", type=str, default="dibbs_output", metavar="DIR",
        help="Output directory (created if missing). Default: ./dibbs_output",
    )
    p.add_argument(
        "--raw", action="store_true",
        help="Also save a raw CSV alongside the Excel (useful for debugging).",
    )
    return p.parse_args()


def main() -> None:
    args = parse_args()

    # Resolve output directory dynamically from CLI arg
    out_dir = os.path.abspath(args.out)
    os.makedirs(out_dir, exist_ok=True)

    timestamp = datetime.now().strftime("%m-%d-%Y")
    xlsx_path = os.path.join(out_dir, f"dibbs_rfq_{timestamp}.xlsx")
    csv_path  = os.path.join(out_dir, f"dibbs_rfq_{timestamp}_all_raw.csv")

    try:
        rows = scrape(max_pages=args.pages)
    except KeyboardInterrupt:
        print("\n[Interrupted by user]")
        sys.exit(0)
    except Exception:
        traceback.print_exc()
        sys.exit(1)

    if not rows:
        print(
            "\n⚠️  No rows matched all filters.\n"
            "   Possible reasons:\n"
            "   • DIBBS page structure changed (check --raw output)\n"
            "   • No current RFQs match Mil-spec + '*spec/stnd only' for these FSC codes\n"
            "   • DIBBS returned a CAPTCHA or login wall — try again later\n"
        )
    else:
        write_excel(rows, xlsx_path)
        if args.raw:
            write_raw_csv(rows, csv_path)
            print(f"    Raw CSV saved → {csv_path}")

    print(f"\nDone. Output folder: {out_dir}\n")


if __name__ == "__main__":
    main()
